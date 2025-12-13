"""Spreadsheet automation module for Excel and CSV files."""

from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from ..core.logger import LoggerMixin


class SpreadsheetModule(LoggerMixin):
    """Handle Excel and CSV file operations."""

    def read_excel(
        self,
        file_path: str,
        sheet_name: Optional[Union[str, int]] = 0,
        header: Optional[int] = 0,
    ) -> pd.DataFrame:
        """Read an Excel file into a DataFrame.

        Args:
            file_path: Path to the Excel file
            sheet_name: Sheet name or index (default: first sheet)
            header: Row to use as header (default: 0)

        Returns:
            DataFrame with the spreadsheet data
        """
        self.logger.info(f"Reading Excel file: {file_path}")
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header)
        self.logger.info(f"Read {len(df)} rows from {file_path}")
        return df

    def read_csv(
        self,
        file_path: str,
        delimiter: str = ",",
        encoding: str = "utf-8",
        header: Optional[int] = 0,
    ) -> pd.DataFrame:
        """Read a CSV file into a DataFrame.

        Args:
            file_path: Path to the CSV file
            delimiter: Column delimiter (default: comma)
            encoding: File encoding (default: utf-8)
            header: Row to use as header (default: 0)

        Returns:
            DataFrame with the CSV data
        """
        self.logger.info(f"Reading CSV file: {file_path}")
        df = pd.read_csv(file_path, delimiter=delimiter, encoding=encoding, header=header)
        self.logger.info(f"Read {len(df)} rows from {file_path}")
        return df

    def write_excel(
        self,
        data: Union[pd.DataFrame, List[Dict[str, Any]]],
        file_path: str,
        sheet_name: str = "Sheet1",
        index: bool = False,
    ) -> str:
        """Write data to an Excel file.

        Args:
            data: DataFrame or list of dicts to write
            file_path: Output file path
            sheet_name: Name for the worksheet
            index: Whether to include the index

        Returns:
            Path to the created file
        """
        if isinstance(data, list):
            data = pd.DataFrame(data)

        self.logger.info(f"Writing {len(data)} rows to Excel: {file_path}")
        data.to_excel(file_path, sheet_name=sheet_name, index=index)
        return file_path

    def write_csv(
        self,
        data: Union[pd.DataFrame, List[Dict[str, Any]]],
        file_path: str,
        index: bool = False,
        encoding: str = "utf-8",
    ) -> str:
        """Write data to a CSV file.

        Args:
            data: DataFrame or list of dicts to write
            file_path: Output file path
            index: Whether to include the index
            encoding: File encoding

        Returns:
            Path to the created file
        """
        if isinstance(data, list):
            data = pd.DataFrame(data)

        self.logger.info(f"Writing {len(data)} rows to CSV: {file_path}")
        data.to_csv(file_path, index=index, encoding=encoding)
        return file_path

    def read(self, file_path: str, **kwargs) -> pd.DataFrame:
        """Auto-detect file type and read accordingly."""
        path = Path(file_path)
        if path.suffix.lower() in [".xlsx", ".xls"]:
            return self.read_excel(file_path, **kwargs)
        elif path.suffix.lower() == ".csv":
            return self.read_csv(file_path, **kwargs)
        else:
            raise ValueError(f"Unsupported file format: {path.suffix}")

    def write(self, data: Union[pd.DataFrame, List[Dict]], file_path: str, **kwargs) -> str:
        """Auto-detect file type and write accordingly."""
        path = Path(file_path)
        if path.suffix.lower() in [".xlsx", ".xls"]:
            return self.write_excel(data, file_path, **kwargs)
        elif path.suffix.lower() == ".csv":
            return self.write_csv(data, file_path, **kwargs)
        else:
            raise ValueError(f"Unsupported file format: {path.suffix}")

    def transform(
        self,
        data: pd.DataFrame,
        operations: List[Dict[str, Any]],
    ) -> pd.DataFrame:
        """Apply a series of transformations to a DataFrame.

        Args:
            data: Input DataFrame
            operations: List of operations, each a dict with 'type' and params

        Supported operations:
            - filter: {'type': 'filter', 'column': 'col', 'condition': 'value'}
            - rename: {'type': 'rename', 'columns': {'old': 'new'}}
            - drop: {'type': 'drop', 'columns': ['col1', 'col2']}
            - sort: {'type': 'sort', 'by': 'column', 'ascending': True}
            - fillna: {'type': 'fillna', 'value': 0}

        Returns:
            Transformed DataFrame
        """
        df = data.copy()

        for op in operations:
            op_type = op.get("type")

            if op_type == "filter":
                column = op["column"]
                condition = op["condition"]
                df = df[df[column] == condition]

            elif op_type == "rename":
                df = df.rename(columns=op["columns"])

            elif op_type == "drop":
                df = df.drop(columns=op["columns"])

            elif op_type == "sort":
                df = df.sort_values(by=op["by"], ascending=op.get("ascending", True))

            elif op_type == "fillna":
                df = df.fillna(op["value"])

            self.logger.debug(f"Applied transformation: {op_type}")

        return df

    def merge(
        self,
        left: pd.DataFrame,
        right: pd.DataFrame,
        on: Union[str, List[str]],
        how: str = "inner",
    ) -> pd.DataFrame:
        """Merge two DataFrames.

        Args:
            left: Left DataFrame
            right: Right DataFrame
            on: Column(s) to merge on
            how: Merge type (inner, left, right, outer)

        Returns:
            Merged DataFrame
        """
        self.logger.info(f"Merging DataFrames on {on} ({how} join)")
        return pd.merge(left, right, on=on, how=how)

    def create_styled_excel(
        self,
        data: Union[pd.DataFrame, List[Dict]],
        file_path: str,
        sheet_name: str = "Sheet1",
        header_style: Optional[Dict] = None,
    ) -> str:
        """Create a styled Excel file with formatting.

        Args:
            data: Data to write
            file_path: Output path
            sheet_name: Sheet name
            header_style: Optional header styling options

        Returns:
            Path to created file
        """
        if isinstance(data, list):
            data = pd.DataFrame(data)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write headers
        for col_idx, column in enumerate(data.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, row in enumerate(data.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, column in enumerate(data.columns, 1):
            max_length = max(
                len(str(column)),
                data[column].astype(str).str.len().max() if len(data) > 0 else 0
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

        wb.save(file_path)
        self.logger.info(f"Created styled Excel file: {file_path}")
        return file_path

    def get_sheet_names(self, file_path: str) -> List[str]:
        """Get list of sheet names from an Excel file."""
        wb = load_workbook(file_path, read_only=True)
        return wb.sheetnames

    def append_to_excel(
        self,
        data: Union[pd.DataFrame, List[Dict]],
        file_path: str,
        sheet_name: str = "Sheet1",
    ) -> str:
        """Append data to an existing Excel file.

        Args:
            data: Data to append
            file_path: Path to existing file
            sheet_name: Sheet to append to

        Returns:
            Path to the file
        """
        if isinstance(data, list):
            data = pd.DataFrame(data)

        if not Path(file_path).exists():
            return self.write_excel(data, file_path, sheet_name)

        existing = self.read_excel(file_path, sheet_name=sheet_name)
        combined = pd.concat([existing, data], ignore_index=True)
        return self.write_excel(combined, file_path, sheet_name)
